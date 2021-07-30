#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
使用第三方库：pip install openpyxl
一般用于处理新版本 Excel(.xlsx)
"""

import openpyxl


# 写入 Excel
# 数据形如：[['张三', '信息与通信工程', '数值分析', 88]]
def write_excel(sheet_name, head, data, path):
    # 实例化一个工作薄对象
    workbook = openpyxl.Workbook()
    # 激活一个 Sheet 表（工作表），并为它设置一个 title
    sheet = workbook.active
    sheet.title = sheet_name

    # data 中添加表头（不需要表头可以不用加）
    data.insert(0, list(head))

    # 开始遍历并插入数据
    # row: 行  col: 列
    for row_index, row_item in enumerate(data):
        for col_index, col_item in enumerate(row_item):
            # 写入单元格
            sheet.cell(row=row_index + 1, column=col_index + 1, value=col_item)

    workbook.save(path)


# 读取 Excel
def read_excel(path, sheet_name):
    # 实例化一个工作薄对象
    workbook = openpyxl.load_workbook(path)
    # 获取指定名字的 Sheet 表
    sheet = workbook[sheet_name]

    # 定义一个数组，存放要输出的数据
    result = []

    # sheet.rows 为表格内的每一行数据
    # 循环获取表格内的每一行数据
    for index, row in enumerate(sheet.rows):
        # 定义一个空的数组用来存放每一行数据单元格的数据
        current_row = []
        for col_index, col_value in enumerate(row):
            # 获取单元格数据 追加到 return_row
            current_row.append(col_value.value)
        # 把每一行数据追加到结果 return_data 中，最后输出
        result.append(current_row)

    return result


if __name__ == "__main__":
    # mock 数据
    sheet_name = '成绩'
    head = ['姓名', '专业', '科目', '成绩']
    data = [
        ['张三', '信息与通信工程', '数值分析', 88],
        ['李四', '物联网工程', '数字信号处理分析', 95],
        ['王华', '电子与通信工程', '模糊数学', 90],
        ['王欢', '通信工程', '机器学习', 89]
    ]
    path = 'student.xlsx'

    # 执行方法
    # write_excel(sheet_name, head, data, path)
    result = read_excel(path, sheet_name)
    print(result)
    # edit_excel()
    # format_excel()
